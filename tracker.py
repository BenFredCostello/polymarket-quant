"""
tracker.py
----------
Persistent trade tracking (JSON) and Excel export for Polymarket bets.
"""

import json
import math
import uuid
import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests

logger = logging.getLogger(__name__)

TRADES_FILE = Path(__file__).parent / "trades.json"
EXCEL_FILE  = Path(__file__).parent / "polymarket_trades.xlsx"
STARTING_BANKROLL = 1_000.0

try:
    from config import RISK_FREE_RATE
except ImportError:
    RISK_FREE_RATE = 0.045


# ── Persistence ───────────────────────────────────────────────────────────────

def _load() -> dict:
    if TRADES_FILE.exists():
        with open(TRADES_FILE) as f:
            return json.load(f)
    return {"starting_bankroll": STARTING_BANKROLL, "trades": []}


def _save(data: dict) -> None:
    with open(TRADES_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)


def record_bet(signal, yes_token: str = "") -> dict:
    """Persist a confirmed bet from a TradeSignal. Returns the stored trade dict."""
    data = _load()
    trade = {
        "id":           str(uuid.uuid4())[:8],
        "date_placed":  datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "condition_id": signal.condition_id,
        "yes_token":    yes_token or getattr(signal, "yes_token", ""),
        "question":     signal.question,
        "asset":        signal.asset,
        "strike":       signal.strike,
        "expiry":       signal.end_date,
        "side":         signal.side,
        "option_type":  signal.option_type,
        "market_prob":  round(signal.market_prob, 4),
        "model_prob_bs": round(signal.model_prob_bs, 4),
        "model_prob_mc": round(signal.model_prob_mc, 4),
        "edge":         round(signal.edge_bs, 4),
        "stake_usd":    round(signal.recommended_stake_usd, 2),
        "status":       "OPEN",
        "outcome":      None,
        "pnl":          None,
        "date_resolved": None,
    }
    data["trades"].append(trade)
    _save(data)
    return trade


# ── Resolution checking ───────────────────────────────────────────────────────

def _extract_yes_price(m: dict) -> float | None:
    """Pull YES price from a Gamma API market dict using multiple fallback methods."""
    # Method 1: tokens array
    for t in m.get("tokens", []):
        if t.get("outcome", "").upper() == "YES":
            try:
                return float(t["price"])
            except (KeyError, ValueError, TypeError):
                pass

    # Method 2: outcomePrices + outcomes JSON strings
    op = m.get("outcomePrices")
    oc = m.get("outcomes")
    if op and oc:
        try:
            prices   = json.loads(op) if isinstance(op, str) else op
            outcomes = json.loads(oc) if isinstance(oc, str) else oc
            for label, price in zip(outcomes, prices):
                if str(label).upper() == "YES":
                    return float(price)
        except Exception:
            pass

    # Method 3: bid/ask midpoint
    bid = m.get("bestBid")
    ask = m.get("bestAsk")
    if bid is not None and ask is not None:
        return (float(bid) + float(ask)) / 2

    return None


def _check_resolution(condition_id: str, yes_token: str) -> tuple[bool, str | None]:
    """
    Returns (is_resolved, outcome) where outcome is "YES", "NO", or None.
    Tries Gamma API first, then CLOB price history as fallback.
    """
    GAMMA = "https://gamma-api.polymarket.com"
    CLOB  = "https://clob.polymarket.com"

    # --- Method 1: Gamma API by conditionId ---
    for param_key in ("conditionId", "id"):
        try:
            resp = requests.get(
                f"{GAMMA}/markets",
                params={param_key: condition_id, "limit": 1},
                timeout=15,
            )
            if resp.status_code == 200:
                data = resp.json()
                markets = data if isinstance(data, list) else data.get("markets", [])
                if markets:
                    m = markets[0]
                    yes_p = _extract_yes_price(m)
                    if yes_p is not None:
                        if yes_p >= 0.99:
                            return True, "YES"
                        if yes_p <= 0.01:
                            return True, "NO"
        except Exception as e:
            logger.debug(f"Gamma resolution check failed ({param_key}={condition_id[:20]}): {e}")

    # --- Method 2: CLOB prices-history (latest data point) ---
    if yes_token:
        try:
            resp = requests.get(
                f"{CLOB}/prices-history",
                params={"market": yes_token, "interval": "max", "fidelity": 1440},
                timeout=15,
            )
            if resp.status_code == 200:
                history = resp.json().get("history", [])
                if history:
                    latest_price = float(max(history, key=lambda h: int(h["t"]))["p"])
                    if latest_price >= 0.99:
                        return True, "YES"
                    if latest_price <= 0.01:
                        return True, "NO"
        except Exception as e:
            logger.debug(f"CLOB resolution check failed for token {yes_token[:20]}: {e}")

    return False, None


def check_resolutions() -> list[dict]:
    """
    Poll the API for all OPEN bets. Marks resolved ones with outcome + P&L.
    Returns a list of newly resolved trade dicts.
    """
    data = _load()
    open_trades = [t for t in data["trades"] if t["status"] == "OPEN"]
    if not open_trades:
        return []

    newly_resolved = []
    for trade in open_trades:
        try:
            is_resolved, resolved_outcome = _check_resolution(
                trade["condition_id"], trade.get("yes_token", "")
            )
        except Exception as e:
            logger.warning(f"Resolution check error for {trade['id']}: {e}")
            continue

        if not is_resolved:
            continue

        stake      = trade["stake_usd"]
        entry_prob = trade["market_prob"]

        won = (
            (trade["side"] == "YES" and resolved_outcome == "YES") or
            (trade["side"] == "NO"  and resolved_outcome == "NO")
        )

        if won:
            if trade["side"] == "YES":
                pnl = stake * (1 - entry_prob) / entry_prob
            else:
                pnl = stake * entry_prob / (1 - entry_prob)
        else:
            pnl = -stake

        trade["status"]        = "WON" if won else "LOST"
        trade["outcome"]       = resolved_outcome
        trade["pnl"]           = round(pnl, 2)
        trade["date_resolved"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        newly_resolved.append(trade)

    _save(data)
    return newly_resolved


def get_portfolio_stats(trades: list[dict], starting_bankroll: float) -> dict:
    resolved     = [t for t in trades if t["status"] in ("WON", "LOST")]
    open_trades  = [t for t in trades if t["status"] == "OPEN"]
    wins         = [t for t in resolved if t["status"] == "WON"]
    losses       = [t for t in resolved if t["status"] == "LOST"]
    total_pnl    = sum(t["pnl"] for t in resolved)

    # Annualised Sharpe: build daily equity curve, compute daily returns,
    # then Sharpe = (mean_daily - rf/365) / std_daily * sqrt(365)
    sharpe = None
    resolved_dated = sorted(
        [t for t in resolved if t.get("date_resolved")],
        key=lambda t: t["date_resolved"],
    )
    if len(resolved_dated) >= 2:
        date_equity: dict[str, float] = {}
        running = starting_bankroll
        for t in resolved_dated:
            running += t["pnl"]
            date_equity[str(t["date_resolved"])[:10]] = running

        dates = sorted(date_equity)
        start = datetime.strptime(dates[0],  "%Y-%m-%d").date()
        end   = datetime.strptime(dates[-1], "%Y-%m-%d").date()
        n_days = (end - start).days + 1

        daily_vals, prev = [], starting_bankroll
        for i in range(n_days):
            d = (start + timedelta(days=i)).strftime("%Y-%m-%d")
            prev = date_equity.get(d, prev)
            daily_vals.append(prev)

        if len(daily_vals) >= 3:
            rets = [(daily_vals[i] - daily_vals[i-1]) / daily_vals[i-1]
                    for i in range(1, len(daily_vals))]
            n    = len(rets)
            mean = sum(rets) / n
            std  = (sum((r - mean) ** 2 for r in rets) / (n - 1)) ** 0.5
            if std > 0:
                daily_rf = RISK_FREE_RATE / 365
                sharpe   = (mean - daily_rf) / std * (365 ** 0.5)

    return {
        "starting_bankroll": starting_bankroll,
        "net_value":         starting_bankroll + total_pnl,
        "total_pnl":         total_pnl,
        "roi":               total_pnl / starting_bankroll if starting_bankroll else 0,
        "win_rate":          len(wins) / len(resolved) if resolved else 0.0,
        "total_bets":        len(trades),
        "resolved_bets":     len(resolved),
        "open_bets":         len(open_trades),
        "wins":              len(wins),
        "losses":            len(losses),
        "at_risk":           sum(t["stake_usd"] for t in open_trades),
        "sharpe":            sharpe,
        "sharpe_n_days":     n_days if sharpe is not None else None,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

# Colour palette
_C = {
    "header_bg":  "2C3E50",
    "accent_bg":  "2980B9",
    "white":      "FFFFFF",
    "light_bg":   "F8F9FA",
    "win_bg":     "D5F5E3",
    "loss_bg":    "FADBD8",
    "open_bg":    "FEF9E7",
    "win_fg":     "1E8449",
    "loss_fg":    "C0392B",
    "muted":      "7F8C8D",
}


def export_excel() -> Path:
    """Generate/refresh the Excel workbook from persisted trade history."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
    except ImportError:
        print("  openpyxl not installed — run: pip install openpyxl")
        return None

    data             = _load()
    trades           = data["trades"]
    starting_bankroll = data.get("starting_bankroll", STARTING_BANKROLL)
    stats            = get_portfolio_stats(trades, starting_bankroll)

    wb = openpyxl.Workbook()

    # ─── helpers ────────────────────────────────────────────────────────────────

    def _fill(color: str):
        return PatternFill("solid", fgColor=color)

    def _font(color=None, bold=False, size=10, italic=False):
        return Font(
            name="Calibri",
            size=size,
            bold=bold,
            italic=italic,
            color=color or "000000",
        )

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _header_cell(cell, text, bg=None, fg=None, size=10, center=False):
        cell.value       = text
        cell.font        = _font(color=fg or _C["white"], bold=True, size=size)
        cell.fill        = _fill(bg or _C["header_bg"])
        cell.alignment   = _align("center" if center else "left")

    # ─── Sheet 1: Portfolio ──────────────────────────────────────────────────────

    ws = wb.active
    ws.title = "Portfolio"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20

    row = 1

    # Title bar
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value     = "POLYMARKET QUANT TRADING"
    c.font      = _font(color=_C["white"], bold=True, size=20)
    c.fill      = _fill(_C["accent_bg"])
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 42
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value     = f"Last updated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    c.font      = _font(color=_C["muted"], size=10, italic=True)
    c.fill      = _fill("EBF5FB")
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 18
    row += 2

    # Summary header
    ws.merge_cells(f"A{row}:B{row}")
    _header_cell(ws[f"A{row}"], "PERFORMANCE SUMMARY", size=11, center=True)
    ws.row_dimensions[row].height = 26
    row += 1

    def _metric(label, value, fmt=None, fg=None):
        nonlocal row
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = _font(bold=False, size=10)
        lc.fill      = _fill(_C["light_bg"])
        lc.alignment = _align()

        vc = ws.cell(row=row, column=2, value=value)
        vc.font      = _font(bold=True, size=10, color=fg)
        vc.fill      = _fill(_C["light_bg"])
        vc.alignment = _align("right")
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[row].height = 20
        row += 1

    pnl_color = _C["win_fg"] if stats["total_pnl"] >= 0 else _C["loss_fg"]
    roi_color = pnl_color

    _metric("Starting Bankroll",       starting_bankroll,      '"$"#,##0.00')
    _metric("Current Portfolio Value", stats["net_value"],    '"$"#,##0.00')
    _metric("Total P&L",               stats["total_pnl"],   '"$"+#,##0.00;"$"-#,##0.00', pnl_color)
    _metric("Total Return",            stats["roi"],          '0.0%;-0.0%',                roi_color)
    _metric("Win Rate",
            stats["win_rate"] if stats["resolved_bets"] else "—",
            '0.0%' if stats["resolved_bets"] else None)
    _metric("Wins / Losses",           f"{stats['wins']} / {stats['losses']}")
    _metric("Total Bets Placed",       stats["total_bets"])
    _metric("Open Positions",          stats["open_bets"])
    _metric("Capital at Risk",         stats["at_risk"],      '"$"#,##0.00')
    _metric("Risk-Free Rate",          RISK_FREE_RATE,        '0.00%')

    sharpe      = stats.get("sharpe")
    sharpe_days = stats.get("sharpe_n_days")
    sharpe_label = (
        f"Sharpe ratio ({sharpe_days}d, annualised)"
        if sharpe_days else "Sharpe Ratio"
    )
    sharpe_val = round(sharpe, 2) if sharpe is not None else "—"
    _metric(sharpe_label, sharpe_val)

    row += 1

    # Equity curve data table
    resolved_sorted = sorted(
        [t for t in trades if t["status"] in ("WON", "LOST") and t.get("date_resolved")],
        key=lambda t: t["date_resolved"],
    )

    ws.merge_cells(f"A{row}:B{row}")
    _header_cell(ws[f"A{row}"], "EQUITY CURVE", size=11, center=True)
    ws.row_dimensions[row].height = 26
    chart_anchor_row = row
    row += 1

    # Column headers
    _header_cell(ws.cell(row=row, column=1), "Date", bg=_C["accent_bg"], center=True, size=10)
    _header_cell(ws.cell(row=row, column=2), "Net Value ($)", bg=_C["accent_bg"], center=True, size=10)
    ws.row_dimensions[row].height = 20
    row += 1

    eq_data_start = row

    # Seed row
    seed_date = resolved_sorted[0]["date_resolved"] if resolved_sorted else datetime.now().strftime("%Y-%m-%d")
    ws.cell(row=row, column=1, value=str(seed_date)[:10])
    ws.cell(row=row, column=2, value=starting_bankroll).number_format = '"$"#,##0.00'
    ws.row_dimensions[row].height = 18
    row += 1

    running = starting_bankroll
    for t in resolved_sorted:
        running += t["pnl"]
        ws.cell(row=row, column=1, value=str(t["date_resolved"])[:10])
        c = ws.cell(row=row, column=2, value=running)
        c.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 18
        row += 1

    eq_data_end = row - 1

    # Chart (only if we have data)
    if eq_data_end > eq_data_start:
        from openpyxl.chart.marker import Marker

        chart = LineChart()
        chart.title  = "Portfolio Net Value Over Time"
        chart.style  = 10
        chart.height = 14
        chart.width  = 26

        values = Reference(ws, min_col=2, min_row=eq_data_start - 1, max_row=eq_data_end)
        dates  = Reference(ws, min_col=1, min_row=eq_data_start,     max_row=eq_data_end)
        from openpyxl.chart.title import Title
        from openpyxl.drawing.text import RichTextProperties, Paragraph, RegularTextRun, CharacterProperties

        chart.add_data(values, titles_from_data=True, from_rows=False)
        chart.set_categories(dates)

        chart.x_axis.delete = False
        chart.y_axis.delete = False

        # Nice y-axis bounds: round data min/max to a clean step, one step of padding each side
        eq_values = [
            ws.cell(row=r, column=2).value
            for r in range(eq_data_start, eq_data_end + 1)
            if ws.cell(row=r, column=2).value is not None
        ]
        if eq_values:
            v_min, v_max = min(eq_values), max(eq_values)
            step = 50
            y_min = math.floor(v_min / step) * step - step
            y_max = math.ceil(v_max  / step) * step + step
            chart.y_axis.scaling.min = float(y_min)
            chart.y_axis.scaling.max = float(y_max)
            chart.y_axis.majorUnit   = float(step)

        t = Title()
        para = Paragraph()
        run = RegularTextRun()
        run.rPr = CharacterProperties(sz=1300, b=True)
        run.t = "Portfolio Net Value Over Time"
        para.r.append(run)
        t.tx.rich.bodyPr = RichTextProperties()
        t.tx.rich.p = [para]
        t.overlay = False
        chart.title = t

        s = chart.series[0]
        s.graphicalProperties.line.solidFill = _C["accent_bg"]
        s.graphicalProperties.line.width     = 25000
        s.marker = Marker(symbol="circle", size=4)
        s.marker.graphicalProperties.solidFill             = _C["accent_bg"]
        s.marker.graphicalProperties.line.solidFill        = _C["accent_bg"]

        chart.legend = None

        ws.add_chart(chart, f"D{chart_anchor_row}")

    # ─── Sheet 2: All Trades ────────────────────────────────────────────────────

    ws2 = wb.create_sheet("All Trades")
    ws2.sheet_view.showGridLines = False

    TRADE_COLS = [
        ("Date Placed",   13),
        ("Expires",       12),
        ("Asset",          7),
        ("Question",      54),
        ("Side",           6),
        ("Mkt Prob",       9),
        ("Model Prob",    10),
        ("Edge",           8),
        ("Stake ($)",     10),
        ("Status",         8),
        ("Outcome",        8),
        ("P&L ($)",       11),
        ("Net Value ($)", 13),
    ]

    for col_idx, (name, width) in enumerate(TRADE_COLS, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
        _header_cell(ws2.cell(row=1, column=col_idx), name, center=True)
    ws2.row_dimensions[1].height = 26
    ws2.freeze_panes = "A2"

    # Sort: most recent first; open positions at top
    all_sorted = sorted(
        trades,
        key=lambda t: (
            0 if t["status"] == "OPEN" else 1,
            -int((t.get("date_resolved") or t.get("date_placed") or "0").replace("-", "") or 0),
        ),
    )

    # Build cumulative P&L map (keyed by trade id, value = running pnl after this trade)
    cumulative: dict[str, float] = {}
    running = 0.0
    for t in resolved_sorted:
        running += t["pnl"]
        cumulative[t["id"]] = running

    current_net = starting_bankroll + running

    for r, trade in enumerate(all_sorted, 2):
        status = trade["status"]
        pnl    = trade.get("pnl")

        if status in ("WON", "LOST"):
            net_val = starting_bankroll + cumulative.get(trade["id"], running)
            row_bg  = _C["win_bg"] if status == "WON" else _C["loss_bg"]
        else:
            net_val = current_net
            row_bg  = _C["open_bg"]

        model_prob = (trade.get("model_prob_bs", 0) + trade.get("model_prob_mc", 0)) / 2

        values = [
            trade.get("date_placed", ""),
            (trade.get("expiry") or "")[:10],
            trade.get("asset", ""),
            trade.get("question", ""),
            trade.get("side", ""),
            trade.get("market_prob"),
            model_prob if model_prob > 0 else None,
            trade.get("edge"),
            trade.get("stake_usd"),
            status,
            trade.get("outcome") or "",
            pnl,
            net_val,
        ]

        for col_idx, val in enumerate(values, 1):
            cell            = ws2.cell(row=r, column=col_idx, value=val)
            cell.font       = _font(size=10)
            cell.fill       = _fill(row_bg)
            cell.alignment  = _align(wrap=(col_idx == 4))

        ws2.cell(row=r, column=6).number_format  = "0.0%"
        ws2.cell(row=r, column=7).number_format  = "0.0%"
        ws2.cell(row=r, column=8).number_format  = "+0.0%;-0.0%"
        ws2.cell(row=r, column=9).number_format  = '"$"#,##0.00'
        ws2.cell(row=r, column=12).number_format = '"$"#,##0.00'
        ws2.cell(row=r, column=13).number_format = '"$"#,##0.00'

        if pnl is not None:
            ws2.cell(row=r, column=12).font = _font(
                color=_C["win_fg"] if pnl >= 0 else _C["loss_fg"],
                bold=True, size=10,
            )

        ws2.row_dimensions[r].height = 20

    # ─── Sheet 3: Open Positions ─────────────────────────────────────────────────

    ws3 = wb.create_sheet("Open Positions")
    ws3.sheet_view.showGridLines = False

    OPEN_COLS = [
        ("Date Placed",  13),
        ("Expires",      12),
        ("Asset",         7),
        ("Question",     54),
        ("Side",          6),
        ("Mkt Prob",      9),
        ("Model Prob",   10),
        ("Edge",          8),
        ("Stake ($)",    10),
        ("Days Left",    10),
    ]

    for col_idx, (name, width) in enumerate(OPEN_COLS, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
        _header_cell(ws3.cell(row=1, column=col_idx), name, bg=_C["accent_bg"], center=True)
    ws3.row_dimensions[1].height = 26
    ws3.freeze_panes = "A2"

    open_trades = [t for t in trades if t["status"] == "OPEN"]

    if not open_trades:
        ws3.merge_cells("A2:J2")
        c            = ws3["A2"]
        c.value      = "No open positions."
        c.font       = _font(color=_C["muted"], italic=True, size=11)
        c.alignment  = _align("center")
    else:
        today = datetime.now().date()
        for r, trade in enumerate(open_trades, 2):
            model_prob = (trade.get("model_prob_bs", 0) + trade.get("model_prob_mc", 0)) / 2

            days_left = ""
            expiry_str = trade.get("expiry") or ""
            if expiry_str:
                try:
                    exp_date  = datetime.fromisoformat(expiry_str.replace("Z", "+00:00")).date()
                    days_left = (exp_date - today).days
                except Exception:
                    pass

            values = [
                trade.get("date_placed", ""),
                expiry_str[:10],
                trade.get("asset", ""),
                trade.get("question", ""),
                trade.get("side", ""),
                trade.get("market_prob"),
                model_prob if model_prob > 0 else None,
                trade.get("edge"),
                trade.get("stake_usd"),
                days_left,
            ]

            for col_idx, val in enumerate(values, 1):
                cell           = ws3.cell(row=r, column=col_idx, value=val)
                cell.font      = _font(size=10)
                cell.fill      = _fill(_C["open_bg"])
                cell.alignment = _align(wrap=(col_idx == 4))

            ws3.cell(row=r, column=6).number_format = "0.0%"
            ws3.cell(row=r, column=7).number_format = "0.0%"
            ws3.cell(row=r, column=8).number_format = "+0.0%;-0.0%"
            ws3.cell(row=r, column=9).number_format = '"$"#,##0.00'

            if isinstance(days_left, int) and days_left <= 3:
                ws3.cell(row=r, column=10).font = _font(color=_C["loss_fg"], bold=True, size=10)

            ws3.row_dimensions[r].height = 20

        # Total at-risk footer
        footer = len(open_trades) + 2
        ws3.cell(row=footer, column=8, value="Total at risk:").font = _font(bold=True, size=10)
        c = ws3.cell(row=footer, column=9, value=stats["at_risk"])
        c.number_format = '"$"#,##0.00'
        c.font          = _font(bold=True, size=10)

    import time
    for attempt in range(5):
        try:
            wb.save(EXCEL_FILE)
            return EXCEL_FILE
        except PermissionError:
            if attempt < 4:
                time.sleep(2)
            else:
                raise
